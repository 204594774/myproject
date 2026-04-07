from flask import Blueprint, request, session, current_app, send_file
from werkzeug.security import generate_password_hash
from app.utils.db import get_db_connection
from app.utils.response import success, fail
from app.utils.auth import user_manage_required
from config import get_config
from flasgger import swag_from
import sqlite3
import io
import os
import time
import uuid
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

config = get_config()
ROLES = config.ROLES

users_bp = Blueprint('users', __name__, url_prefix='/api')

def _user_table_columns(conn):
    return {r[1] for r in conn.execute('PRAGMA table_info(users)').fetchall()}

def _norm(v):
    s = '' if v is None else str(v)
    s = s.replace('\ufeff', '').replace('\u200b', '').replace('\xa0', ' ').replace('　', ' ')
    return s.strip()

def _insert_user(conn, payload):
    cols = _user_table_columns(conn)
    fields = []
    values = []
    for k in [
        'username', 'password', 'role', 'real_name', 'identity_number', 'department', 'college',
        'personal_info', 'teaching_office', 'research_area', 'email', 'phone',
        'grade', 'enrollment_year', 'college_code', 'major_code', 'status'
    ]:
        if k in cols and k in payload:
            fields.append(k)
            values.append(payload.get(k))
    if not fields:
        raise ValueError('用户表结构异常')
    placeholders = ','.join(['?'] * len(fields))
    conn.execute(
        f'INSERT INTO users ({", ".join(fields)}) VALUES ({placeholders})',
        values
    )

@users_bp.route('/users', methods=['GET'])
@user_manage_required
@swag_from({
    'tags': ['用户管理'],
    'summary': '获取用户列表',
    'parameters': [
        {
            'name': 'status',
            'in': 'query',
            'type': 'string',
            'description': '状态过滤 (例如: active, pending)'
        }
    ],
    'responses': {
        200: {'description': '获取成功'}
    }
})
def get_users():
    current_role = session.get('role')
    status_filter = request.args.get('status') # active, pending
    role_filter = (request.args.get('role') or '').strip()
    college_filter = (request.args.get('college') or '').strip()
    keyword = (request.args.get('keyword') or '').strip()
    
    conn = get_db_connection()
    try:
        # 检查现有列，兼容旧库
        cols = [r[1] for r in conn.execute('PRAGMA table_info(users)').fetchall()]
        base_cols = ['id', 'username', 'real_name', 'role']
        opt_cols = []
        for c in ['college', 'department', 'identity_number', 'status', 'email', 'phone', 'teaching_office', 'research_area', 'grade', 'enrollment_year', 'college_code', 'major_code']:
            if c in cols:
                opt_cols.append(c)
        select_cols = ', '.join(base_cols + opt_cols)
        query = f'SELECT {select_cols} FROM users WHERE 1=1'
        params = []
        
        if status_filter and 'status' in cols:
            query += ' AND status = ?'
            params.append(status_filter)
        if role_filter:
            query += ' AND role = ?'
            params.append(role_filter)
        
        if current_role == ROLES['COLLEGE_APPROVER']:
            user_college = session.get('college', '')
            # 记录调试日志以便排查
            current_app.logger.info(f"Admin College: {user_college}, Role: {current_role}")
            
            if user_college:
                # 极致模糊匹配：只要包含关键名称即可
                short_name = user_college.replace('（人工智能学院）', '').replace('(人工智能学院)', '').strip()
                query += ' AND (college LIKE ? OR college LIKE ? OR college = ?)'
                params.append(f"%{short_name}%")
                params.append(f"%计算机学院%")
                params.append(user_college)
            
            # 允许查看本学院所有角色，不加 role 限制以防漏掉
        
        elif current_role in [ROLES['SCHOOL_APPROVER'], ROLES['SYSTEM_ADMIN'], ROLES['PROJECT_ADMIN']]:
            # 校级管理员可以看到所有人，不加任何过滤
            pass

        if college_filter and current_role != ROLES['COLLEGE_APPROVER']:
            query += ' AND college LIKE ?'
            params.append(f"%{college_filter}%")

        if keyword:
            query += ' AND (username LIKE ? OR real_name LIKE ? OR identity_number LIKE ?)'
            kw = f"%{keyword}%"
            params.extend([kw, kw, kw])
            
        users = conn.execute(query, params).fetchall()
        
        res = []
        for row in users:
            d = dict(row)
            for k in ['college', 'department', 'identity_number', 'status', 'email', 'phone', 'teaching_office', 'research_area', 'grade', 'enrollment_year', 'college_code', 'major_code']:
                if k not in d:
                    d[k] = ''
                # 如果 status 为空（可能是旧数据或未设置），默认为 active
                if k == 'status' and not d[k]:
                    d[k] = 'active'
            res.append(d)
        return success(data=res)
    except Exception as e:
        return fail(str(e), 500)

@users_bp.route('/users', methods=['POST'])
@user_manage_required
def create_user():
    current_role = session.get('role')
    data = request.json or {}
    new_role = _norm(data.get('role'))
    
    if current_role == ROLES['PROJECT_ADMIN']:
        if new_role not in [ROLES['TEACHER'], ROLES['STUDENT'], ROLES['JUDGE']]:
            return fail('只能创建导师/学生/评审专家账号', 403)
    if current_role == ROLES['COLLEGE_APPROVER']:
        if new_role not in [ROLES['TEACHER'], ROLES['STUDENT']]:
            return fail('只能创建导师或学生账号', 403)
            
    conn = get_db_connection()
    try:
        college = _norm(data.get('college'))
        if current_role == ROLES['COLLEGE_APPROVER']:
            college = session.get('college', '')
            if not college:
                return fail('学院管理员缺少学院信息，无法创建用户', 400)

        identity = _norm(data.get('identity_number'))
        username = _norm(data.get('username'))
        if new_role in [ROLES['STUDENT'], ROLES['TEACHER']]:
            if identity:
                username = identity
            elif username:
                identity = username
        if not username:
            return fail('账号不能为空', 400)

        hashed_password = generate_password_hash(data.get('password', '123456'))
        payload = {
            'username': username,
            'password': hashed_password,
            'role': new_role,
            'real_name': _norm(data.get('real_name')),
            'identity_number': identity,
            'department': _norm(data.get('department')),
            'college': college,
            'personal_info': _norm(data.get('personal_info')),
            'teaching_office': _norm(data.get('teaching_office')),
            'research_area': _norm(data.get('research_area')),
            'email': _norm(data.get('email')),
            'phone': _norm(data.get('phone')),
            'grade': _norm(data.get('grade')),
            'enrollment_year': _norm(data.get('enrollment_year')),
            'college_code': _norm(data.get('college_code')),
            'major_code': _norm(data.get('major_code')),
            'status': _norm(data.get('status')) or 'active'
        }
        _insert_user(conn, payload)
        conn.commit()
        return success(message='用户创建成功')
    except sqlite3.IntegrityError:
        return fail('用户名已存在', 400)
    except Exception as e:
        return fail(str(e), 500)

@users_bp.route('/users/<int:uid>', methods=['PUT'])
@user_manage_required
def update_user(uid):
    current_role = session.get('role')
    data = request.json
    target_role = data.get('role')
    
    conn = get_db_connection()
    # Check target user exists
    target_user = conn.execute('SELECT * FROM users WHERE id = ?', (uid,)).fetchone()
    if not target_user:
        return fail('用户不存在', 404)
        
    # Permission Check
    if current_role == ROLES['PROJECT_ADMIN']:
        if target_user['role'] == ROLES['SYSTEM_ADMIN']:
            return fail('无权修改系统管理员', 403)
        if target_role and target_role not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('只能设置为学生或导师角色', 403)
    elif current_role == ROLES['COLLEGE_APPROVER']:
        if (target_user['college'] or '') != (session.get('college', '') or ''):
            return fail('无权修改非本院用户', 403)
        if target_user['role'] not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('无权修改该用户', 403)
        if 'college' in data and data.get('college') != session.get('college', ''):
            return fail('学院管理员不能修改用户学院', 403)
        if target_role and target_role not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('只能设置为学生或导师角色', 403)

    try:
        # Build Update Query
        fields = []
        values = []
        
        update_fields = [
            'real_name', 'identity_number', 'college', 'department', 'teaching_office', 'research_area',
            'email', 'phone', 'grade', 'enrollment_year', 'college_code', 'major_code',
            'role', 'status'
        ]
        for f in update_fields:
            if f in data:
                fields.append(f'{f} = ?')
                values.append(data[f])
                
        if 'password' in data and data['password']:
            fields.append('password = ?')
            values.append(generate_password_hash(data['password']))
            
        if not fields:
            return success(message='无变更')
            
        values.append(uid)
        conn.execute(f'UPDATE users SET {", ".join(fields)} WHERE id = ?', values)
        conn.commit()
        return success(message='更新成功')
    except Exception as e:
        return fail(str(e), 500)

@users_bp.route('/users/<int:uid>', methods=['DELETE'])
@user_manage_required
def delete_user(uid):
    current_role = session.get('role')
    conn = get_db_connection()
    target_user = conn.execute('SELECT * FROM users WHERE id = ?', (uid,)).fetchone()
    if not target_user:
        return fail('用户不存在', 404)
        
    if current_role == ROLES['PROJECT_ADMIN']:
        if target_user['role'] == ROLES['SYSTEM_ADMIN']:
            return fail('无权删除系统管理员', 403)
    elif current_role == ROLES['COLLEGE_APPROVER']:
        if (target_user['college'] or '') != (session.get('college', '') or ''):
            return fail('无权删除非本院用户', 403)
        if target_user['role'] not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('无权删除该用户', 403)
            
    try:
        conn.execute('DELETE FROM users WHERE id = ?', (uid,))
        conn.commit()
        return success(message='删除成功')
    except Exception as e:
        return fail(str(e), 500)

@users_bp.route('/users/<int:uid>/approve', methods=['PUT'])
@user_manage_required
def approve_user(uid):
    data = request.json
    action = data.get('action') # approve, reject
    
    conn = get_db_connection()
    current_role = session.get('role')
    if current_role == ROLES['COLLEGE_APPROVER']:
        target_user = conn.execute('SELECT id, college, role FROM users WHERE id = ?', (uid,)).fetchone()
        if not target_user:
            return fail('用户不存在', 404)
        if (target_user['college'] or '') != (session.get('college', '') or ''):
            return fail('无权审核非本院用户', 403)
        if target_user['role'] not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('无权审核该用户', 403)

    if action == 'approve':
        conn.execute('UPDATE users SET status = "active" WHERE id = ?', (uid,))
    elif action == 'reject':
        conn.execute('UPDATE users SET status = "rejected" WHERE id = ?', (uid,))
    
    conn.commit()
    return success(message='操作成功')


@users_bp.route('/users/<int:uid>/reset_password', methods=['POST'])
@user_manage_required
def reset_password(uid):
    conn = get_db_connection()
    current_role = session.get('role')
    target_user = conn.execute('SELECT id, role, college FROM users WHERE id = ?', (uid,)).fetchone()
    if not target_user:
        return fail('用户不存在', 404)

    if current_role == ROLES['PROJECT_ADMIN'] and target_user['role'] == ROLES['SYSTEM_ADMIN']:
        return fail('无权重置系统管理员密码', 403)
    if current_role == ROLES['COLLEGE_APPROVER']:
        if (target_user['college'] or '') != (session.get('college', '') or ''):
            return fail('无权操作非本院用户', 403)
        if target_user['role'] not in [ROLES['STUDENT'], ROLES['TEACHER']]:
            return fail('无权操作该用户', 403)

    try:
        conn.execute(
            'UPDATE users SET password = ? WHERE id = ?',
            (generate_password_hash('123456'), uid)
        )
        conn.commit()
        return success(message='已重置为 123456')
    except Exception as e:
        return fail(str(e), 500)


def _workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _auto_width(ws, headers):
    for i, h in enumerate(headers, start=1):
        w = max(12, min(40, len(str(h or '')) + 6))
        ws.column_dimensions[get_column_letter(i)].width = w

def _save_bytes_to_upload(content, suffix):
    folder = current_app.config.get('UPLOAD_FOLDER') or os.path.join(os.getcwd(), 'static', 'uploads')
    os.makedirs(folder, exist_ok=True)
    filename = f"{int(time.time())}_{uuid.uuid4().hex[:8]}{suffix}"
    path = os.path.join(folder, filename)
    with open(path, 'wb') as f:
        f.write(content)
    return f"/static/uploads/{filename}"


@users_bp.route('/users/export', methods=['GET'])
@user_manage_required
def export_users():
    current_role = session.get('role')
    status_filter = request.args.get('status')
    role_filter = (request.args.get('role') or '').strip()
    college_filter = (request.args.get('college') or '').strip()
    keyword = (request.args.get('keyword') or '').strip()

    conn = get_db_connection()
    cols = _user_table_columns(conn)
    base_cols = ['id', 'username', 'real_name', 'role']
    opt_cols = []
    for c in ['college', 'department', 'identity_number', 'status', 'email', 'phone', 'teaching_office', 'research_area', 'grade', 'enrollment_year', 'college_code', 'major_code']:
        if c in cols:
            opt_cols.append(c)
    select_cols = base_cols + opt_cols
    query = f"SELECT {', '.join(select_cols)} FROM users WHERE 1=1"
    params = []
    if status_filter and 'status' in cols:
        query += ' AND status = ?'
        params.append(status_filter)
    if role_filter:
        query += ' AND role = ?'
        params.append(role_filter)
    if current_role == ROLES['COLLEGE_APPROVER']:
        user_college = session.get('college', '')
        if user_college:
            short_name = user_college.replace('（人工智能学院）', '').replace('(人工智能学院)', '').strip()
            query += ' AND (college LIKE ? OR college LIKE ? OR college = ?)'
            params.append(f"%{short_name}%")
            params.append(f"%计算机学院%")
            params.append(user_college)
    else:
        if college_filter:
            query += ' AND college LIKE ?'
            params.append(f"%{college_filter}%")
    if keyword:
        kw = f"%{keyword}%"
        query += ' AND (username LIKE ? OR real_name LIKE ? OR identity_number LIKE ?)'
        params.extend([kw, kw, kw])
    rows = conn.execute(query, params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'users'
    headers = ['ID', '账号(工号/学号)', '姓名', '角色', '学院', '专业/职称', '学号/工号', '年级', '入学年份', '邮箱', '电话', '状态']
    ws.append(headers)
    for r in rows:
        d = dict(r)
        ws.append([
            d.get('id', ''),
            d.get('username', ''),
            d.get('real_name', ''),
            d.get('role', ''),
            d.get('college', ''),
            d.get('department', ''),
            d.get('identity_number', ''),
            d.get('grade', ''),
            d.get('enrollment_year', ''),
            d.get('email', ''),
            d.get('phone', ''),
            d.get('status', '')
        ])
    _auto_width(ws, headers)
    download_name = f"账号列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(_workbook_bytes(wb), as_attachment=True, download_name=download_name)


@users_bp.route('/admin/import/students/template', methods=['GET'])
@user_manage_required
def download_student_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'students'
    headers = ['姓名', '学号', '学院', '专业', '年级', '入学年份', '学院代码', '专业代码']
    ws.append(headers)
    _auto_width(ws, headers)
    return send_file(_workbook_bytes(wb), as_attachment=True, download_name='学生批量导入模板.xlsx')


@users_bp.route('/admin/import/teachers/template', methods=['GET'])
@user_manage_required
def download_teacher_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'teachers'
    headers = ['姓名', '工号', '学院', '职称', '邮箱', '联系电话']
    ws.append(headers)
    _auto_width(ws, headers)
    return send_file(_workbook_bytes(wb), as_attachment=True, download_name='教师批量导入模板.xlsx')


def _max_student_serial(conn, prefix8):
    row = conn.execute(
        "SELECT MAX(CAST(SUBSTR(username, 9, 4) AS INTEGER)) AS mx FROM users WHERE username GLOB ?",
        (f"{prefix8}[0-9][0-9][0-9][0-9]",)
    ).fetchone()
    mx = row['mx'] if row else None
    return int(mx) if mx is not None else None

def _max_teacher_suffix(conn, prefix5):
    row = conn.execute(
        "SELECT MAX(CAST(SUBSTR(username, 6, 4) AS INTEGER)) AS mx FROM users WHERE username GLOB ?",
        (f"{prefix5}[0-9][0-9][0-9][0-9]",)
    ).fetchone()
    mx = row['mx'] if row else None
    return int(mx) if mx is not None else None

def _read_xlsx_rows(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    header_row = 1
    header = []
    scan_max = min(20, int(getattr(ws, 'max_row', 1) or 1))
    for rno in range(1, scan_max + 1):
        vals = []
        for cell in ws[rno]:
            vals.append(_norm(cell.value))
        s = {v for v in vals if v}
        if ('姓名' in s) and ('学院' in s) and (('学号' in s) or ('工号' in s)):
            header_row = rno
            header = vals
            break
    if not header:
        for cell in ws[1]:
            header.append(_norm(cell.value))
    idx = {h: i for i, h in enumerate(header) if h}
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not r or all(_norm(x) == '' for x in r):
            continue
        rows.append((idx, r))
    return rows


@users_bp.route('/admin/import/students', methods=['POST'])
@user_manage_required
def import_students():
    if 'file' not in request.files:
        return fail('未上传文件', 400)
    file = request.files['file']
    if not file or not file.filename:
        return fail('文件名为空', 400)
    if not str(file.filename).lower().endswith('.xlsx'):
        return fail('仅支持 .xlsx', 400)

    conn = get_db_connection()
    current_role = session.get('role')
    session_college = _norm(session.get('college', ''))
    success_count = 0
    failures = []
    try:
        for idx_map, row in _read_xlsx_rows(file):
            def getv(k):
                i = idx_map.get(k)
                return row[i] if i is not None and i < len(row) else ''

            name = _norm(getv('姓名'))
            student_id = _norm(getv('学号'))
            college = _norm(getv('学院'))
            major = _norm(getv('专业'))
            grade = _norm(getv('年级'))
            enrollment_year = _norm(getv('入学年份'))
            college_code = _norm(getv('学院代码'))
            major_code = _norm(getv('专业代码'))

            if not name or not college or not enrollment_year:
                failures.append({'row': row, 'reason': '必填字段缺失（姓名/学院/入学年份）'})
                continue

            if current_role == ROLES['COLLEGE_APPROVER']:
                if not session_college:
                    failures.append({'row': row, 'reason': '学院管理员缺少学院信息'})
                    continue
                college = session_college

            if enrollment_year:
                try:
                    enrollment_year = str(int(float(enrollment_year)))
                except Exception:
                    enrollment_year = _norm(enrollment_year)

            if student_id:
                username = student_id
            else:
                if not college_code or not major_code:
                    failures.append({'row': row, 'reason': '学号为空时需填写学院代码/专业代码'})
                    continue
                cc = college_code.zfill(2) if college_code.isdigit() else college_code
                mc = major_code.zfill(2) if major_code.isdigit() else major_code
                prefix8 = f"{enrollment_year}{cc}{mc}"
                mx = _max_student_serial(conn, prefix8)
                next_serial = 1 if mx is None else (mx + 1)
                if next_serial > 9999:
                    failures.append({'row': row, 'reason': '学号流水号已满'})
                    continue
                username = f"{prefix8}{next_serial:04d}"
                student_id = username
                college_code = cc
                major_code = mc

            exists = conn.execute('SELECT 1 FROM users WHERE username = ?', (username,)).fetchone()
            if exists:
                failures.append({'row': row, 'reason': f'学号重复：{username}'})
                continue

            payload = {
                'username': username,
                'password': generate_password_hash('123456'),
                'role': ROLES['STUDENT'],
                'real_name': name,
                'identity_number': student_id,
                'college': college,
                'department': major,
                'grade': grade,
                'enrollment_year': enrollment_year,
                'college_code': college_code,
                'major_code': major_code,
                'status': 'active'
            }
            _insert_user(conn, payload)
            success_count += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        return fail(str(e), 500)

    fail_count = len(failures)
    fail_url = ''
    if failures:
        wb = Workbook()
        ws = wb.active
        ws.title = 'failures'
        headers = ['失败原因', '姓名', '学号', '学院', '专业', '年级', '入学年份', '学院代码', '专业代码']
        ws.append(headers)
        for f in failures:
            r = f.get('row') or []
            def cell(i):
                return _norm(r[i]) if i < len(r) else ''
            ws.append([f.get('reason', ''), cell(0), cell(1), cell(2), cell(3), cell(4), cell(5), cell(6), cell(7)])
        _auto_width(ws, headers)
        url = _save_bytes_to_upload(_workbook_bytes(wb).getvalue(), '.xlsx')
        fail_url = url

    return success(data={'success': success_count, 'failed': fail_count, 'fail_url': fail_url})


@users_bp.route('/admin/import/teachers', methods=['POST'])
@user_manage_required
def import_teachers():
    if 'file' not in request.files:
        return fail('未上传文件', 400)
    file = request.files['file']
    if not file or not file.filename:
        return fail('文件名为空', 400)
    if not str(file.filename).lower().endswith('.xlsx'):
        return fail('仅支持 .xlsx', 400)

    conn = get_db_connection()
    current_role = session.get('role')
    session_college = _norm(session.get('college', ''))
    success_count = 0
    failures = []
    try:
        for idx_map, row in _read_xlsx_rows(file):
            def getv(k):
                i = idx_map.get(k)
                return row[i] if i is not None and i < len(row) else ''

            name = _norm(getv('姓名'))
            work_id = _norm(getv('工号'))
            college = _norm(getv('学院'))
            title = _norm(getv('职称'))
            email = _norm(getv('邮箱'))
            phone = _norm(getv('联系电话'))

            if not name or not college:
                failures.append({'row': row, 'reason': '必填字段缺失（姓名/学院）'})
                continue

            if current_role == ROLES['COLLEGE_APPROVER']:
                if not session_college:
                    failures.append({'row': row, 'reason': '学院管理员缺少学院信息'})
                    continue
                college = session_college

            if not work_id:
                prefix = 'T2026'
                mx = _max_teacher_suffix(conn, prefix)
                next_suffix = 0 if mx is None else (mx + 1)
                if next_suffix > 9999:
                    failures.append({'row': row, 'reason': '工号流水号已满'})
                    continue
                work_id = f"{prefix}{next_suffix:04d}"

            exists = conn.execute('SELECT 1 FROM users WHERE username = ?', (work_id,)).fetchone()
            if exists:
                failures.append({'row': row, 'reason': f'工号重复：{work_id}'})
                continue

            payload = {
                'username': work_id,
                'password': generate_password_hash('123456'),
                'role': ROLES['TEACHER'],
                'real_name': name,
                'identity_number': work_id,
                'college': college,
                'department': title,
                'email': email,
                'phone': phone,
                'status': 'active'
            }
            _insert_user(conn, payload)
            success_count += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        return fail(str(e), 500)

    fail_count = len(failures)
    fail_url = ''
    if failures:
        wb = Workbook()
        ws = wb.active
        ws.title = 'failures'
        headers = ['失败原因', '姓名', '工号', '学院', '职称', '邮箱', '联系电话']
        ws.append(headers)
        for f in failures:
            r = f.get('row') or []
            def cell(i):
                return _norm(r[i]) if i < len(r) else ''
            ws.append([f.get('reason', ''), cell(0), cell(1), cell(2), cell(3), cell(4), cell(5)])
        _auto_width(ws, headers)
        url = _save_bytes_to_upload(_workbook_bytes(wb).getvalue(), '.xlsx')
        fail_url = url

    return success(data={'success': success_count, 'failed': fail_count, 'fail_url': fail_url})


@users_bp.route('/admin/sync-from-cms', methods=['POST'])
@user_manage_required
def sync_from_cms():
    # TODO: 对接学校教务系统
    mock = {
        'students': [
            {'real_name': '张三', 'student_id': '202600010001', 'college': '计算机学院', 'major': '软件工程', 'grade': '2026级', 'enrollment_year': '2026', 'college_code': '01', 'major_code': '01'}
        ],
        'teachers': [
            {'real_name': '李四', 'work_id': 'T20260000', 'college': '计算机学院', 'title': '讲师', 'email': 'lisi@school.edu.cn', 'phone': '13800000000'}
        ]
    }
    return success(data=mock)
